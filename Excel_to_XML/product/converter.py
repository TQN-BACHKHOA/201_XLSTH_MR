from openpyxl import load_workbook
from yattag import Doc, indent

wb = load_workbook("NY_baby_names.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

doc.asis(xml_header)
doc.asis(xml_schema)

with tag('Babies'):
    # Use ws.max_row for all rows
    for row in ws.iter_rows(min_row=2, max_row=100, min_col=1, max_col=5):
        row = [cell.value for cell in row]
        with tag("Baby"):
            with tag("Name"):
                text(row[2]) #tất cả các hàng của cột thứ 2
            with tag("Gender"):
                text(row[1])
            with tag("year"):
                text(row[0])
            with tag("count"):
                text(row[3])
            with tag("rank"):
                text(row[4])

result = indent(
    doc.getvalue(),
    indentation = '    ',
    indent_text = True
)

with open("baby_names.xml", "w") as f:
    f.write(result)
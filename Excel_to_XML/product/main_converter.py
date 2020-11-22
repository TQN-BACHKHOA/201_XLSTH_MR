from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from yattag import Doc, indent
import datetime

wb = load_workbook("data_simplified.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

doc.asis(xml_header)
doc.asis(xml_schema)

current_date = datetime.datetime.now()
temp_date = datetime.datetime(2020, 11, 22)

with tag('Cac_Benh_Nhan'):
    # Use ws.max_row for all rows
    for col in ws.iter_cols(min_col=2, max_col=3, min_row=2, max_row=41):
        col = [cell.value for cell in col]
        count = 0
        with tag("Benh_Nhan"):
            with tag("MA_LK", klass='code', type='string_100'):
                if col[count] == None or len(str(col[count])) > 100:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("STT", klass='number', type='int_10'):
                if col[count] == None or len(str(col[count])) > 10:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_BN", klass='code', type='string_100'):
                if col[count] == None or len(col[count]) > 100:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("HO_TEN", klass='detail', type='string_255'):
                if col[count] == None or len(col[count]) > 255:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("NGAY_SINH", klass='time', type='year_month_day'):
                if type(col[count]) is datetime.datetime:
                    temp_date = col[count]
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("GIOI_TINH", klass='detail', type='selection_1'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Nam":
                    text('1')
                elif col[count] == "Nữ":
                    text('2')
                elif col[count] == "Chưa xác định":
                    text('3')
                count += 1
            with tag("DIA_CHI", klass='detail', type='string_1024'):
                if col[count] == None or len(col[count]) > 1024:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_THE", klass='code', type='string_n'):
                if col[count] == None:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_DKBD", klass='code', type='string_n'):
                if col[count] == None:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1 
            with tag("GT_THE_TU", klass='time', type='year_month_day'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("GT_THE_DEN", klass='time', type='year_month_day'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("MIEN_CUNG_CT", klass='time', type='year_month_day'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("TEN_BENH", klass='detail', type='string_n'):
                if  col[count] == None:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_BENH", klass='code', type='string_15'):
                if col[count] == None or len(col[count]) >= 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_BENHKHAC", klass='code', type='string_255'):
                if col[count] == None or len(col[count]) >= 255:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_LYDO_VVIEN", klass='detail', type='selection_1'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Đúng tuyến":
                    text("1")
                elif col[count] == "Cấp cứu":
                    text("2")
                elif col[count] == "Trái tuyến":
                    text("3")
                elif col[count] == "Thông tuyến":
                    text("4")
                count += 1
            with tag("MA_NOI_CHUYEN", klass='code', type='string_5'):
                if col[count] == None or len(col[count]) > 5:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_TAI_NAN", klass='code', type='int_1'):
                if type(col[count]) == int and len(str(col[count])) == 1:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("NGAY_VAO", klass='time', type='year_month_day_hour_minute'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d%H%M"))
                else:
                    text("NODATA")
                count += 1
            with tag("NGAY_RA", klass='time', type='year_month_day_hour_minute'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d%H%M"))
                else:
                    text("NODATA")
                count += 1
            with tag("SO_NGAY_DTRI", klass='time', type='int_3'):
                if type(col[count]) == int and len(str(col[count])) <= 3:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("KET_QUA_DTRI", klass='detail', type='selection_1'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Khỏi":
                    text("1")
                elif col[count] == "Đỡ":
                    text("2")
                elif col[count] == "Không thay đổi":
                    text("3")
                elif col[count] == "Nặng hơn":
                    text("4")
                elif col[count] == "Tử vong":
                    text("5")
                count += 1
            with tag("TINH_TRANG_RV", klass='detail', type='selection_1'):
                if col[count] == None:
                    text("NODATA")
                if col[count] == "Ra viện":
                    text('1')
                elif col[count] == "Chuyển viện":
                    text('2')
                elif col[count] == "Trốn viện":
                    text('3')
                elif col[count] == "Xin ra viện":
                    text('4')
                count += 1
            with tag("NGAY_TTOAN", klass='time', type='year_month_day'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d%H%M"))
                else:
                    text("NODATA")
                count += 1
            with tag("T_THUOC", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("T_VTYT", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("T_TONGCHI", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("T_BNTT", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("T_BNCCT", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("T_BHTT", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("T_NGUONKHAC", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("T_NGOAIDS", klass='money', type='float_15_decimal_2'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("NAM_QT", klass='time', type='year'):
                if type(col[count]) == int and len(str(col[count])) == 4:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("THANG_QT", klass='time', type='month'):
                if type(col[count]) == int and len(str(col[count])) == 2:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("MA_LOAI_KCB", klass='code', type='selection_1'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Khám bệnh":
                    text('1')
                elif col[count] == "Điều trị ngoại trú":
                    text('2')
                elif col[count] == "Điều trị nội trú":
                    text('3')
                count += 1
            with tag("MA_KHOA", klass='code', type='string_15'):
                if col[count] == None or len(col[count]) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_CSKCB", klass='code', type='string_5'):
                if col[count] == None or len(col[count]) > 5:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("MA_KHUVUC", klass='code', type='selection_2'):
                if col[count] == None:
                    text("NODATA")
                if col[count] == "K1":
                    text('1')
                elif col[count] == "K2":
                    text('2')
                elif col[count] == "K3":
                    text('3')
                count += 1
            with tag("MA_PTTT_QT", klass='code', type='string_255'):
                if col[count] == None or len(col[count]) > 255:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("CAN_NANG", klass='detail', type='float_5_decimal_2'):
                if (current_date - temp_date).days <= 365:
                    if type(col[count]) == float and len(str(col[count])) <= 5:
                        text(col[count])
                        age = (current_date - temp_date).days
                        print(current_date)
                        print(temp_date)
                        print(age)
                else:
                    text("NODATA")

result = indent(
    doc.getvalue(),
    indentation = '    '
)

with open("patient_names.xml", "w") as f:
    f.write(result)
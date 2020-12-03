from openpyxl import load_workbook
from yattag import Doc, indent
import datetime

wb = load_workbook("data_simplified.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
#xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

doc.asis(xml_header)
#doc.asis(xml_schema)

current_date = datetime.datetime.now()
temp_date = datetime.datetime(2020, 11, 22)
#...

with tag('Cac_Benh_Nhan'):
    # Use ws.max_row for all rows
    for col in ws.iter_cols(min_col=2, max_col=3, min_row=2, max_row=41):
        col = [cell.value for cell in col]
        count = 0
        with tag("Benh_Nhan"):
            with tag("node", klass='code', type='string_100', id='MA_LK'):
                if col[count] == None or len(str(col[count])) > 100:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='int_10', id='STT'):
                if col[count] == None or len(str(col[count])) > 10:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='code', type='string_100', id='MA_BN'):
                if col[count] == None or len(col[count]) > 100:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='code', type='string_255', id='HO_TEN'):
                if col[count] == None or len(col[count]) > 255:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='time', type='year_month_day', id='NGAY_SINH'):
                if type(col[count]) is datetime.datetime:
                    temp_date = col[count]
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='detail', type='selection_1', id='GIOI_TINH'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Nam":
                    text('1')
                elif col[count] == "Nữ":
                    text('2')
                elif col[count] == "Chưa xác định":
                    text('3')
                count += 1
            with tag("node", klass='code', type='string_1024', id='DIA_CHI'):
                if col[count] == None or len(col[count]) > 1024:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='code', type='string_n', id='MA_THE'):
                if col[count] == None:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='code', type='string_n', id='MA_DKBD'):
                if col[count] == None:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1 
            with tag("node", klass='time', type='year_month_day', id='GT_THE_TU'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='time', type='year_month_day', id='GT_THE_DEN'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='time', type='year_month_day', id='MIEN_CUNG_CT'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d"))
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='code', type='string_n', id='TEN_BENH'):
                if  col[count] == None:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='code', type='string_15', id='MA_BENH'):
                if col[count] == None or len(col[count]) >= 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='code', type='string_255', id='MA_BENHKHAC'):
                if col[count] == None or len(col[count]) >= 255:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='detail', type='selection_1', id='MA_LYDO_VVIEN'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Đúng tuyến":
                    text("1")
                elif col[count] == "Cấp cứu":
                    text("2")
                elif col[count] == "Trái tuyến":
                    text("3")
                elif col[count] == "Thông tuyến":
                    text("4")
                count += 1
            with tag("node", klass='code', type='string_5', id='MA_NOI_CHUYEN'):
                if col[count] == None or len(col[count]) > 5:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='int_1', id='MA_TAI_NAN'):
                if type(col[count]) == int and len(str(col[count])) == 1:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='time', type='year_month_day_hour_minute', id='NGAY_VAO'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d%H%M"))
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='time', type='year_month_day_hour_minute', id='NGAY_RA'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d%H%M"))
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='number', type='int_3', id='SO_NGAY_DTRI'):
                if type(col[count]) == int and len(str(col[count])) <= 3:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='detail', type='selection_1', id='KET_QUA_DTRI'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Khỏi":
                    text("1")
                elif col[count] == "Đỡ":
                    text("2")
                elif col[count] == "Không thay đổi":
                    text("3")
                elif col[count] == "Nặng hơn":
                    text("4")
                elif col[count] == "Tử vong":
                    text("5")
                count += 1
            with tag("node", klass='detail', type='selection_1', id='TINH_TRANG_RV'):
                if col[count] == None:
                    text("NODATA")
                if col[count] == "Ra viện":
                    text('1')
                elif col[count] == "Chuyển viện":
                    text('2')
                elif col[count] == "Trốn viện":
                    text('3')
                elif col[count] == "Xin ra viện":
                    text('4')
                count += 1
            with tag("node", klass='time', type='year_month_day', id='NGAY_TTOAN'):
                if type(col[count]) is datetime.datetime:
                    text(col[count].strftime("%Y%m%d%H%M"))
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_THUOC'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_VTYT'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_TONGCHI'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_BNTT'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_BNCCT'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_BHTT'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_NGUONKHAC'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_15_decimal_2', id='T_NGOAIDS'):
                if col[count] == None or len(str(col[count])) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='time', type='year', id='NAM_QT'):
                if type(col[count]) == int and len(str(col[count])) == 4:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='time', type='month', id='THANG_QT'):
                if type(col[count]) == int and len(str(col[count])) == 2:
                    text(col[count])
                else:
                    text("NODATA")
                count += 1
            with tag("node", klass='detail', type='selection_1', id='MA_LOAI_KCB'):
                if col[count] == None:
                    text("NODATA")
                elif col[count] == "Khám bệnh":
                    text('1')
                elif col[count] == "Điều trị ngoại trú":
                    text('2')
                elif col[count] == "Điều trị nội trú":
                    text('3')
                count += 1
            with tag("node", klass='code', type='string_15', id='MA_KHOA'):
                if col[count] == None or len(col[count]) > 15:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='code', type='string_5', id='MA_CSKCB'):
                if col[count] == None or len(col[count]) > 5:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='detail', type='selection_2', id='MA_KHUVUC'):
                if col[count] == None:
                    text("NODATA")
                if col[count] == "K1":
                    text('1')
                elif col[count] == "K2":
                    text('2')
                elif col[count] == "K3":
                    text('3')
                count += 1
            with tag("node", klass='code', type='string_255', id='MA_PTTT_QT'):
                if col[count] == None or len(col[count]) > 255:
                    text("NODATA")
                else:
                    text(col[count])
                count += 1
            with tag("node", klass='number', type='float_5_decimal_2', id='CAN_NANG'):
                if (current_date - temp_date).days <= 365:
                    if type(col[count]) == float and len(str(col[count])) <= 5:
                        text(col[count])
                        age = (current_date - temp_date).days
                        print(current_date)
                        print(temp_date)
                        print(age)
                else:
                    text("NODATA")

result = indent(
    doc.getvalue(),
    indentation = '    '
)

with open("output/patient_names.xml", "w") as f:
    f.write(result)
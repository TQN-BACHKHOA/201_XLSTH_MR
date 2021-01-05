import os
from openpyxl import load_workbook
from yattag import Doc, indent
from queue import PriorityQueue
from os import path

BEGIN_ROW = 1
END_ROW = 143
BEGIN_COL = 2
END_COL = 3

if not path.exists("output"):
    os.mkdir("output")
path1 = 'output'

wb = load_workbook("data_vn.xlsx")
ws = wb.worksheets[0]

XMLNode_stack = []
XML_list = PriorityQueue()
XML_list2 = []
lineCounter = 0
offset = 0

class XMLNode(object):
    def __init__(self, start=None, end=None, text=None):
        self.start = start
        self.end = end
        self.text = text
    def printNodes(self, data, tag, doc):
        global lineCounter
        global offset
        while ((lineCounter+offset) == XML_list2[lineCounter][0]) and (lineCounter < len(XML_list2)):
            if isinstance(XML_list2[lineCounter][3], XMLNode):
                with tag(XML_list2[lineCounter][2]):
                    lineCounter += 1
                    XML_list2[lineCounter-1][3].printNodes(data, tag, doc)
            elif isinstance(XML_list2[lineCounter][3], str):
                if data[lineCounter+offset] is None:
                    tmp_value = ''
                else:
                    tmp_value = data[lineCounter+offset]
                doc.stag(XML_list2[lineCounter][2], '', value=tmp_value)
                lineCounter += 1
        offset += 1

def scan_excelFile():
    for col in ws.iter_cols(min_col=1, max_col=1, min_row=BEGIN_ROW, max_row=END_ROW):
        col = [cell.value for cell in col]
        for _line, _data in enumerate(col):
            _code = _data.split('_')
            if _code[0] == "NODEOPEN":
                _node = XMLNode(start=_line, text=_code[1].lower())
                XMLNode_stack.append(_node)
            elif _code[0] == "NODECLOSE":
                _node = XMLNode_stack.pop()
                _node.end = _line
                XML_list.put((_node.start, _node.end, _node.text, _node))
            else:
                XML_list.put((_line, -1, _code[1].lower(), _data))
    while not XML_list.empty():
	    XML_list2.append(XML_list.get())
    XML_list2.append((-1, -1, -1, -1))
    return _node

def export_xmlFile(rootNode):
    global lineCounter
    global offset
    for column in range(BEGIN_COL, END_COL + 1):
        for col in ws.iter_cols(min_col=column, max_col=column, min_row=BEGIN_ROW, max_row=END_ROW):
            col = [cell.value for cell in col]
            lineCounter = 0
            offset = 0

            # Create Yattag doc, tag and text objects
            doc, tag, text, line = Doc().ttl()
            xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
            # xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'
            doc.asis(xml_header)
            # doc.asis(xml_schema)
            rootNode.printNodes(col, tag, doc)

            file = 'patient' + str(column-1) + '.xml'
            with open(os.path.join(path1, file), 'w') as fp: 
                pass

            result = indent(
            doc.getvalue(),
            indentation = '    '
            )

            with open(os.path.join(path1, file), 'w') as f:
                f.write(result)
        
patientNode = scan_excelFile()
for i in XML_list2:
    print(i)
export_xmlFile(patientNode)

# result = indent(
#     doc.getvalue(),
#     indentation = '    '
# )

# with open("output/patient2.xml", "w") as f:
#     f.write(result)